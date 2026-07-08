#!/usr/bin/env python3
"""
BOM 成本对齐脚本 v1.4
输入: 原始BOM(22列, 机械/电气至少一张) + 采购成本表 → 输出: 手工版格式的成本核算Excel

用法:
  python3 bom-cost-align.py <项目号> [选项]
  python3 bom-cost-align.py 11015
  python3 bom-cost-align.py 11719 --dry-run
"""
import argparse, os, re, sys
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font

# ── 样式 ──
thin = Side(style='thin'); medium = Side(style='medium')
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_border = Border(left=thin, right=thin, top=thin, bottom=medium)
COMP_FILL = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
HDR_FILL = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
COL_WIDTHS = {'A':38,'B':22,'C':28,'D':12,'E':8,'F':12,'G':14,'H':14,'I':10,'J':14}
HDRS = ['件号','零件名称','型号规格','说明','数量','品牌','单价','价格','组件数量','合计总价']

# ── 工具 ──
def hdr_qty(pn):
    m = re.search(r'×(\d+)', pn)
    return int(m.group(1)) if m else 1

def is_comp(pn): return '（' in pn and '）' in pn
def is_elec(pn): return bool(re.match(r'^E\d+-\d+$', pn))

def row_texts(row_data, limit=8):
    return [str(v or '').strip() for v in row_data[:limit]]

def is_elec_section_row(row_data):
    vals = row_texts(row_data)
    nonempty = [v for v in vals if v]
    if len(nonempty) != 1:
        return None
    title = nonempty[0]
    if is_elec(title) or is_comp(title):
        return None
    if title.startswith('件号') or 'Detail No.' in title:
        return None
    return title

def sheet_or_none(wb, name):
    return wb[name] if name in wb.sheetnames else None

# ── 采购加载 ──
def norm_header(v):
    return re.sub(r'\s+', '', str(v or '').strip())

def to_float(v, default=0):
    if v is None or v == '':
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default

def detect_header(ws, required_any, max_scan=20):
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(max_scan, ws.max_row), values_only=True), start=1):
        headers = [norm_header(v) for v in row]
        if any(any(key in h for h in headers) for key in required_any):
            return ridx, headers
    return None, []

def idx_first(headers, *candidates):
    for cand in candidates:
        for i, h in enumerate(headers):
            if cand in h:
                return i
    return None

def load_standard_cost_sheet(ws):
    hr, headers = detect_header(ws, ['图号/零件号', '零件编号', '件号'])
    if not hr:
        return {}
    pn_i = idx_first(headers, '图号/零件号', '零件编号', '件号')
    amount_i = idx_first(headers, '含税金额')
    if pn_i is None or amount_i is None:
        return {}
    cm = {}
    for r in ws.iter_rows(min_row=hr + 1, values_only=True):
        pn = str(r[pn_i] or '').strip() if pn_i < len(r) else ''
        if not pn or pn.startswith('图号') or is_comp(pn):
            continue
        cm[pn] = to_float(r[amount_i] if amount_i < len(r) else 0)
    return cm

def load_costed_bom_sheet(ws):
    hr, headers = detect_header(ws, ['零件编号', '件号'])
    if not hr:
        return {}
    pn_i = idx_first(headers, '零件编号', '件号')
    total_i = idx_first(headers, '总价')
    total_qty_i = idx_first(headers, '总数')
    qty_i = idx_first(headers, '数量')
    rate_i = idx_first(headers, '汇率')
    unit_i = idx_first(headers, '物料含税单价', '含税单价')
    process_i = idx_first(headers, '加工费用总价')
    if pn_i is None:
        return {}
    cm = {}
    for r in ws.iter_rows(min_row=hr + 1, values_only=True):
        pn = str(r[pn_i] or '').strip() if pn_i < len(r) else ''
        if not pn or pn.startswith('件号') or pn.startswith('零件编号') or is_comp(pn):
            continue
        amount = 0
        if total_i is not None and total_i < len(r):
            amount = to_float(r[total_i])
        if amount <= 0 and unit_i is not None:
            qty_src = total_qty_i if total_qty_i is not None else qty_i
            qty = to_float(r[qty_src] if qty_src is not None and qty_src < len(r) else 0)
            rate = to_float(r[rate_i] if rate_i is not None and rate_i < len(r) else 1, 1)
            unit = to_float(r[unit_i] if unit_i < len(r) else 0)
            process = to_float(r[process_i] if process_i is not None and process_i < len(r) else 0)
            amount = qty * rate * unit + process
        cm.setdefault(pn, []).append(amount)
    return cm

def load_cost(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    cm = {}
    sources = []
    for s in wb.sheetnames:
        ws = wb[s]
        part_map = load_standard_cost_sheet(ws)
        if not part_map:
            part_map = load_costed_bom_sheet(ws)
        if part_map:
            for pn, amount in part_map.items():
                if isinstance(amount, list):
                    existing = cm.get(pn)
                    if isinstance(existing, list):
                        existing.extend(amount)
                    elif existing is None:
                        cm[pn] = amount
                    else:
                        cm[pn] = [existing] + amount
                else:
                    cm[pn] = amount
            sources.append(f'{s}:{len(part_map)}')
    wb.close()
    if not cm:
        raise SystemExit(f'❌ 采购成本表缺少可识别的零件级价格明细: {path}')
    print(f'   采购来源: {", ".join(sources)}')
    return cm

def lookup_cost(cost_map, pn):
    amounts = cost_map.get(pn)
    if not amounts:
        return 0
    if isinstance(amounts, list):
        return amounts.pop(0) if amounts else 0
    return amounts

# ── BOM 解析 (统一22列布局) ──
# 机械: [0]件号 [1]名称 [2]类别 [3]规格 [4]物料代码 [5]说明 [6]数量 [7]品牌 ...
# 电气: [0]件号 [1]名称 [2]规格 [3]型号? [4]? [5]说明 [6]数量 [7]品牌 ...

def parse_mech(ws, cost_map):
    rows, stack = [], []
    for row_data in ws.iter_rows(min_row=2, values_only=True):
        pn = str(row_data[0] or '').strip()
        if not pn: continue

        if is_comp(pn):
            code = pn.split('（')[0]
            name = str(row_data[1] or '').strip()
            qty = hdr_qty(pn)
            while stack and code.count('-') <= stack[-1][0].count('-'):
                stack.pop()
            stack.append((code, name or code, qty))
            rows.append(('hdr', {'pn':pn, 'code':code, 'name':name, 'qty':qty}))
            continue

        if 'P' not in pn: continue
        try:
            name = str(row_data[1] or '').strip()
            spec = str(row_data[3] or '').strip()
            desc = str(row_data[5] or '').strip()
            brand = str(row_data[7] or '').strip()
            qty = float(row_data[6]) if row_data[6] else 0
        except (ValueError, TypeError):
            continue
        if qty <= 0: continue

        ct = lookup_cost(cost_map, pn)
        up = ct / qty if qty > 0 and ct > 0 else 0
        rows.append(('part', {
            'pn':pn,'name':name,'spec':spec,'desc':desc,
            'qty':qty,'brand':brand,'unit':up,'total':ct,
            'header': stack[-1][0] if stack else ''
        }))
    return rows

def parse_elec(ws, cost_map):
    rows, cur_section = [], ''
    for row_data in ws.iter_rows(min_row=2, values_only=True):
        pn = str(row_data[0] or '').strip()
        if not pn: continue

        section = is_elec_section_row(row_data)
        if section:
            cur_section = section
            rows.append(('sec', {'pn':section, 'code':section, 'name':section, 'qty':''}))
            continue
        if is_comp(pn): continue
        if not is_elec(pn): continue
        try:
            name = str(row_data[1] or '').strip()
            spec = str(row_data[2] or '').strip()
            desc = str(row_data[5] or '').strip()
            brand = str(row_data[7] or '').strip()
            qty = float(row_data[6]) if row_data[6] else 0
        except (ValueError, TypeError):
            continue
        if qty <= 0: continue
        ct = lookup_cost(cost_map, pn)
        up = ct / qty if qty > 0 and ct > 0 else 0
        rows.append(('part', {
            'pn':pn,'name':name,'spec':spec,'desc':desc,
            'qty':qty,'brand':brand,'unit':up,'total':ct,
            'header': cur_section
        }))
    return rows

def calc_subtotals(rows):
    ct, cur = {}, None
    for rt, d in rows:
        if rt in ('hdr', 'sec'):
            cur = d['code']; ct[cur] = 0
        elif rt == 'part' and cur:
            ct[cur] = ct.get(cur, 0) + d['total']
    return ct

def stats(rows):
    parts = sum(1 for rt,_ in rows if rt == 'part')
    matched = sum(1 for rt,d in rows if rt == 'part' and d['total'] > 0)
    total = sum(d['total'] for rt,d in rows if rt == 'part')
    return parts, matched, total

def fmt_stats(label, rows):
    if rows is None:
        return f'{label}: SKIP'
    parts, matched, total = stats(rows)
    pct = matched / parts * 100 if parts else 0
    return f'{label}: {parts}p/{matched}m ({pct:.0f}%) ¥{total:,.2f}'

# ── 写入 ──
def write_sheet(ws_out, data, comp_total):
    for c, h in enumerate(HDRS, 1):
        cell = ws_out.cell(row=1, column=c, value=h)
        cell.fill = HDR_FILL; cell.border = header_border; cell.font = Font(bold=True)

    r = 2
    for rt, d in data:
        if rt in ('hdr', 'sec'):
            for c in range(1, 11):
                ws_out.cell(row=r, column=c).fill = COMP_FILL
                ws_out.cell(row=r, column=c).border = thin_border
            ws_out.cell(row=r, column=1, value=d['pn'])
            ws_out.cell(row=r, column=2, value=d['name'])
            if rt == 'hdr':
                ws_out.cell(row=r, column=9, value=d['qty'])
                ws_out.cell(row=r, column=10, value=round(comp_total.get(d['code'], 0) * d['qty'], 2))
            else:
                ws_out.cell(row=r, column=10, value=round(comp_total.get(d['code'], 0), 2))
        else:
            for c in range(1, 11):
                ws_out.cell(row=r, column=c).border = thin_border
            ws_out.cell(row=r, column=1, value=d['pn'])
            ws_out.cell(row=r, column=2, value=d['name'])
            ws_out.cell(row=r, column=3, value=d['spec'])
            ws_out.cell(row=r, column=4, value=d['desc'])
            ws_out.cell(row=r, column=5, value=d['qty'])
            ws_out.cell(row=r, column=6, value=d['brand'])
            ws_out.cell(row=r, column=7, value=round(d['unit'], 4))
            ws_out.cell(row=r, column=8, value=round(d['total'], 4))
        r += 1
    for col, w in COL_WIDTHS.items():
        ws_out.column_dimensions[col].width = w

def find_files(d, proj):
    bf = cf = None
    bom_candidates, cost_candidates = [], []
    for f in os.listdir(d):
        fp = os.path.join(d, f)
        if not os.path.isfile(fp): continue
        if not f.endswith('.xlsx') or f.startswith('~$') or proj not in f:
            continue
        if f == f'{proj}.xlsx':
            bom_candidates.insert(0, fp)
        elif 'BOM' in f.upper():
            bom_candidates.append(fp)
        elif any(k in f for k in ('采购成本', '成本核算', '成本')):
            cost_candidates.append(fp)
    bf = bom_candidates[0] if bom_candidates else None
    cf = cost_candidates[0] if cost_candidates else None
    return bf, cf

def main():
    p = argparse.ArgumentParser(description='BOM成本对齐')
    p.add_argument('project'); p.add_argument('--base'); p.add_argument('--bom'); p.add_argument('--cost')
    p.add_argument('--out'); p.add_argument('--dry-run', action='store_true')
    a = p.parse_args()

    base = a.base or os.environ.get('BOM_BASE') or os.getcwd()
    odir = a.out or os.environ.get('BOM_OUT') or os.path.join(os.getcwd(), 'bom-align-output')
    pd = os.path.join(base, a.project)

    bp, cp = (a.bom, a.cost) if a.bom and a.cost else find_files(pd, a.project)
    if not bp: print(f'❌ {pd}/{a.project}.xlsx'); sys.exit(1)
    if not cp: print(f'❌ 采购: {pd}'); sys.exit(1)

    print(f'📦 {a.project}  BOM: {os.path.basename(bp)}  采购: {os.path.basename(cp)}')
    cost = load_cost(cp)
    print(f'   采购项: {len(cost)}')

    wb = openpyxl.load_workbook(bp, data_only=True)
    mech_ws = sheet_or_none(wb, '机械BOM表')
    elec_ws = sheet_or_none(wb, '电气BOM表')
    if mech_ws is None and elec_ws is None:
        available = ', '.join(wb.sheetnames)
        wb.close()
        print(f'❌ 原始BOM缺少 机械BOM表 / 电气BOM表；当前sheet: {available}')
        sys.exit(1)
    mech = parse_mech(mech_ws, cost) if mech_ws is not None else None
    elec = parse_elec(elec_ws, cost) if elec_ws is not None else None
    wb.close()

    mtot = calc_subtotals(mech) if mech is not None else {}
    etot = calc_subtotals(elec) if elec is not None else {}
    mt = stats(mech)[2] if mech is not None else 0
    et = stats(elec)[2] if elec is not None else 0
    es = sum(1 for rt,_ in elec if rt=='sec') if elec is not None else 0

    print(f'   {fmt_stats("机械", mech)}  '
          f'{fmt_stats("电气", elec)}  '
          f'电气栏目: {es if elec is not None else "SKIP"}  合计: ¥{mt+et:,.2f}')

    if a.dry_run: print('🔍 --dry-run'); return

    os.makedirs(os.path.join(odir, a.project), exist_ok=True)
    op = os.path.join(odir, a.project, f'{a.project}_BOM成本核算_自动生成.xlsx')
    wb2 = openpyxl.Workbook(); wb2.remove(wb2.active)
    if mech is not None:
        write_sheet(wb2.create_sheet('机械BOM'), mech, mtot)
    if elec is not None:
        write_sheet(wb2.create_sheet('电气BOM'), elec, etot)
    wb2.save(op)
    print(f'✅ {op}')

if __name__ == '__main__':
    main()
