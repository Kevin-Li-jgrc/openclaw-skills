from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADERS = [
    "Rule ID", "类别", "检查项目", "来源", "严重等级", "自动结论", "证据", "影响", "建议",
    "复核状态", "复核备注", "负责人", "完成日期",
]


def json_text(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def render_xlsx(review: dict[str, Any], output_path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "审查汇总"
    checklist = workbook.create_sheet("自检清单")
    sources = workbook.create_sheet("来源与版本")

    project = review["project"]
    package = review["rule_package"]
    summary_rows = [
        ("项目名称", project.get("name")),
        ("项目路径", project.get("root")),
        ("项目指纹", project.get("fingerprint")),
        ("执行状态", review.get("execution_state")),
        ("总体结论", review.get("overall_status")),
        ("规则版本", package.get("catalog_version")),
        ("规则生效日", package.get("effective_date")),
        ("规则总数", len(review["rules"])),
    ]
    for row in summary_rows:
        summary.append(row)
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 80
    for cell in summary["A"]:
        cell.font = Font(bold=True)

    checklist.append(HEADERS)
    for rule in review["rules"]:
        checklist.append(
            [
                rule["rule_id"],
                rule["rule_id"].split("-")[1],
                rule["title"],
                json_text(rule.get("sources", [])),
                rule["severity"],
                rule["status"],
                json_text(rule.get("evidence", [])),
                rule.get("impact", ""),
                rule.get("recommendation", ""),
                "",
                "", "", "",
            ]
        )

    style_header(checklist)
    checklist.freeze_panes = "A2"
    checklist.auto_filter.ref = f"A1:M{checklist.max_row}"
    checklist.row_dimensions[1].height = 28
    widths = [18, 10, 42, 34, 10, 22, 48, 28, 32, 14, 30, 14, 14]
    for index, width in enumerate(widths, start=1):
        checklist.column_dimensions[get_column_letter(index)].width = width
    for row in checklist.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    validation = DataValidation(type="list", formula1='"未复核,确认问题,误报,已修复,不适用"', allow_blank=True)
    checklist.add_data_validation(validation)
    validation.add(f"J2:J{max(checklist.max_row, 2)}")
    checklist.conditional_formatting.add(
        f"F2:F{checklist.max_row}",
        FormulaRule(formula=["$F2=\"FAIL\""], fill=PatternFill("solid", fgColor="FFC7CE")),
    )

    sources.append(["字段", "值"])
    sources.append(["规则包", "vgage-program-reviewer"])
    sources.append(["规则版本", package.get("catalog_version")])
    sources.append(["生效日期", package.get("effective_date")])
    sources.append(["项目指纹", project.get("fingerprint")])
    style_header(sources)
    sources.column_dimensions["A"].width = 22
    sources.column_dimensions["B"].width = 90

    workbook.properties.title = "VGAGE Pro 程序审查清单"
    workbook.properties.subject = f"Project {project.get('name', '')}"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = output_path.with_name(output_path.stem + ".tmp.xlsx")
    workbook.save(temporary)
    os.replace(temporary, output_path)
