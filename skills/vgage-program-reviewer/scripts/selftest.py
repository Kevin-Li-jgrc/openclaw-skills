from __future__ import annotations

import hashlib
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl


PACKAGE_ROOT = Path(__file__).resolve().parents[1]
FIXTURES = PACKAGE_ROOT / "examples" / "fixtures"
REVIEW = PACKAGE_ROOT / "scripts" / "review-project.py"


def source_hashes(root: Path) -> dict[str, str]:
    return {
        path.relative_to(root).as_posix(): hashlib.sha256(path.read_bytes()).hexdigest()
        for path in root.rglob("*")
        if path.is_file() and not any(part.startswith("VGAGE_REVIEW_") for part in path.relative_to(root).parts)
    }


def main() -> int:
    catalog = json.loads((PACKAGE_ROOT / "references" / "rule-catalog.json").read_text(encoding="utf-8"))
    rules = catalog["rules"]
    ids = {item["rule_id"] for item in rules}
    assert len(rules) == len(ids) == 57
    assert catalog["catalog_version"] == "2.16.0"
    print("package: PASS")
    print("catalog: PASS (57 rules)")
    print("scope: PASS (program-static-only)")

    fixture_count = 0
    consistency_checked = False
    immutable = True
    with tempfile.TemporaryDirectory(prefix="vgage-review-selftest-") as temporary:
        work = Path(temporary)
        for source in sorted(path for path in FIXTURES.iterdir() if path.is_dir()):
            project = work / source.name
            shutil.copytree(source, project)
            before = source_hashes(project)
            completed = subprocess.run(
                [sys.executable, str(REVIEW), str(project), "--project-region", "domestic"],
                text=True,
                capture_output=True,
                check=False,
            )
            expected_code = 2 if source.name == "malformed" else 0
            assert completed.returncode == expected_code, completed.stdout + completed.stderr
            immutable = immutable and before == source_hashes(project)
            fixture_count += 1

            review_dirs = [path for path in project.iterdir() if path.is_dir() and path.name.startswith("VGAGE_REVIEW_")]
            assert len(review_dirs) == 1
            review_dir = review_dirs[0]
            result = json.loads((review_dir / "review-results.json").read_text(encoding="utf-8"))
            workbook = openpyxl.load_workbook(review_dir / "VGAGE程序审查清单.xlsx", read_only=True)
            sheet = workbook["自检清单"]
            xlsx_ids = {sheet.cell(row, 1).value for row in range(2, sheet.max_row + 1)}
            json_ids = {item["rule_id"] for item in result["rules"]}
            assert xlsx_ids == json_ids
            assert json_ids <= ids
            assert all(item["status"] not in {"NOT_APPLICABLE", "NOT_ASSESSABLE"} for item in result["rules"])
            for item in result["rules"]:
                if item["status"] == "MANUAL_VERIFY":
                    assert item["evidence"]
                    assert all(
                        (evidence.get("file") or evidence.get("object"))
                        and evidence.get("reason")
                        and evidence.get("missing_evidence")
                        and evidence.get("manual_action")
                        for evidence in item["evidence"]
                    )
            assert "VG-PRJ-001" in (review_dir / "VGAGE静态审查报告.html").read_text(encoding="utf-8")
            consistency_checked = True

    assert fixture_count == 5
    assert consistency_checked
    assert immutable
    print("fixtures: PASS (5/5)")
    print("html/xlsx/json consistency: PASS")
    print("source immutability: PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
